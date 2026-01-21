from __future__ import annotations
import csv, io
from werkzeug.security import generate_password_hash
from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from .. import db
from ..models import User, Skill, StudentSkill, Attempt, Question

bp = Blueprint("chairman", __name__)

def _ensure_admin():
    if current_user.role != "chairman":
        flash("Chairman access only.", "error")
        return False
    return True

@bp.get("/dashboard")
@login_required
def dashboard():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    teachers = User.query.filter_by(role="teacher").all()
    students = User.query.filter_by(role="student").all()
    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    attempts = Attempt.query.filter(Attempt.finished_at.isnot(None)).all()

    perf = []
    for t in teachers:
        t_attempts = [a for a in attempts if a.teacher_id == t.id]
        avg = round(100 * (sum([a.score or 0 for a in t_attempts]) / len(t_attempts)), 2) if t_attempts else 0.0
        perf.append({"teacher": t, "attempts": len(t_attempts), "avg": avg})
    perf.sort(key=lambda x: x["avg"], reverse=True)

    sperf = []
    for s in students:
        s_attempts = [a for a in attempts if a.student_id == s.id]
        avg = round(100 * (sum([a.score or 0 for a in s_attempts]) / len(s_attempts)), 2) if s_attempts else 0.0
        sperf.append({"student": s, "attempts": len(s_attempts), "avg": avg})
    sperf.sort(key=lambda x: x["avg"])

    return render_template("chairman_dashboard.html", teachers=teachers, students=students, skills=skills, perf=perf, sperf=sperf)

@bp.get("/users")
@login_required
def users():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    teachers = User.query.filter_by(role="teacher").order_by(User.name.asc()).all()
    students = User.query.filter_by(role="student").order_by(User.name.asc()).all()
    return render_template("chairman_users.html", teachers=teachers, students=students)

@bp.post("/users/add_teacher")
@login_required
def add_teacher():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    tid = (request.form.get("tid") or "").strip()
    name = (request.form.get("name") or "").strip()
    email = (request.form.get("email") or "").strip()
    pin = (request.form.get("pin") or "1234").strip()

    if not tid or not name:
        flash("Teacher ID and name required.", "error")
        return redirect(url_for("chairman.users"))

    if User.query.filter_by(id=tid).first():
        flash("ID already exists.", "error")
        return redirect(url_for("chairman.users"))

    db.session.add(User(
        id=tid,
        role="teacher",
        name=name,
        email=email or None,
        pin_hash=generate_password_hash(pin),
    ))
    db.session.commit()
    flash("Teacher added.", "ok")
    return redirect(url_for("chairman.users"))

@bp.post("/users/import_students")
@login_required
def import_students():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    f = request.files.get("csv_file")
    if not f or not f.filename:
        flash("Upload CSV file.", "error")
        return redirect(url_for("chairman.users"))

    content = f.read().decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(content))
    created, updated = 0, 0

    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()

    for row in reader:
        sid = (row.get("student_id") or "").strip()
        name = (row.get("name") or "").strip()
        pin = (row.get("pin") or "1234").strip()
        teacher_id = (row.get("teacher_id") or "").strip()

        if not sid or not name:
            continue

        u = User.query.filter_by(id=sid, role="student").first()
        if not u:
            u = User(
                id=sid,
                role="student",
                name=name,
                teacher_id=teacher_id or None,
                pin_hash=generate_password_hash(pin),
            )
            db.session.add(u)
            created += 1
        else:
            u.name = name
            u.teacher_id = teacher_id or u.teacher_id
            updated += 1

        db.session.flush()

        for sk in skills:
            perm = StudentSkill.query.filter_by(student_id=u.id, skill_id=sk.id).first()
            if not perm:
                allowed = (sk.order_index == 1)
                db.session.add(StudentSkill(student_id=u.id, skill_id=sk.id, allowed=allowed))

    db.session.commit()
    flash(f"Students imported. Created: {created}, Updated: {updated}.", "ok")
    return redirect(url_for("chairman.users"))

@bp.get("/skills")
@login_required
def skills():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    skills = Skill.query.order_by(Skill.order_index.asc()).all()
    return render_template("chairman_skills.html", skills=skills)

@bp.post("/skills/add")
@login_required
def add_skill():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    name = (request.form.get("name") or "").strip()
    order_index = int(request.form.get("order_index") or "0")
    duration_min = int(request.form.get("duration_min") or "0") or None

    pass_pct_raw = request.form.get("pass_pct")
    pass_pct = int(pass_pct_raw) if pass_pct_raw and str(pass_pct_raw).strip() else None

    if not name:
        flash("Skill name required.", "error")
        return redirect(url_for("chairman.skills"))

    sk = Skill(name=name, order_index=order_index, duration_min=duration_min, pass_pct=pass_pct, is_active=True)
    db.session.add(sk)
    db.session.commit()

    students = User.query.filter_by(role="student").all()
    for stu in students:
        db.session.add(StudentSkill(student_id=stu.id, skill_id=sk.id, allowed=False))
    db.session.commit()

    flash("Skill added.", "ok")
    return redirect(url_for("chairman.skills"))

@bp.get("/question_tool")
@login_required
def question_tool():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    questions = Question.query.order_by(Question.id.desc()).limit(500).all()
    return render_template("question_tool.html", skills=skills, questions=questions, role=current_user.role)

@bp.post("/question_tool/add")
@login_required
def add_question():
    # reuse teacher handler logic
    from ..routes.teacher import add_question as teacher_add_question
    if not _ensure_admin():
        return redirect(url_for('auth.home'))
    return teacher_add_question()


@bp.get("/media")
@login_required
def media_library():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))
    import os
    from flask import current_app
    media_dir = current_app.config["MEDIA_DIR"]
    os.makedirs(media_dir, exist_ok=True)
    files = [n for n in sorted(os.listdir(media_dir)) if os.path.isfile(os.path.join(media_dir, n))]
    return render_template("chairman_media.html", files=files)

@bp.post("/media/upload")
@login_required
def media_upload():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))
    import os
    from flask import current_app
    f = request.files.get("file")
    if not f or not f.filename:
        flash("Select a file.", "error")
        return redirect(url_for("chairman.media_library"))
    ext = f.filename.rsplit(".",1)[-1].lower() if "." in f.filename else ""
    if ext not in current_app.config["MEDIA_ALLOWED_EXT"]:
        flash("Media type not allowed.", "error")
        return redirect(url_for("chairman.media_library"))
    safe = "".join([c if c.isalnum() or c in "._-" else "_" for c in f.filename]).strip("_") or ("media."+ext)
    media_dir = current_app.config["MEDIA_DIR"]
    os.makedirs(media_dir, exist_ok=True)
    f.save(os.path.join(media_dir, safe))
    flash("Uploaded.", "ok")
    return redirect(url_for("chairman.media_library"))

@bp.get("/attempts")
@login_required
def attempts():
    if not _ensure_admin():
        return redirect(url_for('auth.home'))

    attempts = Attempt.query.filter(Attempt.finished_at.isnot(None)).order_by(Attempt.finished_at.desc()).limit(500).all()
    return render_template("chairman_attempts.html", attempts=attempts)

@bp.get("/question_import")
@login_required
def question_import():
    if not _ensure_admin():
        return redirect(url_for("auth.home"))
    skills = Skill.query.filter_by(is_active=True).order_by(Skill.order_index.asc()).all()
    return render_template("chairman_question_import.html", skills=skills)

@bp.post("/question_import/upload")
@login_required
def question_import_upload():
    if not _ensure_admin():
        return redirect(url_for("auth.home"))

    import io
    import csv
    import json
    from openpyxl import load_workbook

    f = request.files.get("file")
    default_skill_id = int(request.form.get("default_skill_id") or "0") or None

    if not f or not f.filename:
        flash("Upload a CSV/XLSX file.", "error")
        return redirect(url_for("chairman.question_import"))

    name = f.filename.lower()
    rows = []

    if name.endswith(".csv"):
        content = f.read().decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    elif name.endswith(".xlsx"):
        wb = load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
            rows.append(d)
    else:
        flash("Only .csv or .xlsx supported.", "error")
        return redirect(url_for("chairman.question_import"))

    created = 0
    skipped = 0

    for row in rows:
        skill_id = row.get("skill_id")
        skill_name = row.get("skill_name")
        qtype = (row.get("qtype") or "").strip()
        prompt = (row.get("prompt") or "").strip()
        if not prompt or not qtype:
            skipped += 1
            continue

        sid = None
        try:
            sid = int(skill_id) if skill_id not in (None, "", "None") else None
        except Exception:
            sid = None
        if sid is None and default_skill_id:
            sid = default_skill_id
        if sid is None and skill_name:
            sk = Skill.query.filter_by(name=str(skill_name).strip()).first()
            sid = sk.id if sk else None
        if sid is None:
            skipped += 1
            continue

        options_raw = (row.get("options") or "").strip()
        options_json = None
        if options_raw:
            opts = [x.strip() for x in str(options_raw).split("|") if x.strip()]
            options_json = json.dumps(opts, ensure_ascii=False)

        ans_raw = row.get("answer")
        answer_json = None
        try:
            if qtype == "mcq_multi":
                answer_json = json.dumps([int(x.strip()) for x in str(ans_raw).split(",") if x.strip()], ensure_ascii=False)
            elif qtype == "short_text":
                answer_json = json.dumps(str(ans_raw or ""), ensure_ascii=False)
            else:
                answer_json = json.dumps(int(ans_raw), ensure_ascii=False) if ans_raw not in (None, "", "None") else None
        except Exception:
            answer_json = json.dumps(str(ans_raw or ""), ensure_ascii=False) if qtype == "short_text" else None

        meta = (row.get("meta_json") or row.get("meta") or "")
        meta_json = None
        if meta not in (None, "", "None"):
            m = str(meta).strip()
            meta_json = m if (m.startswith("{") or m.startswith("[")) else json.dumps(m, ensure_ascii=False)

        q = Question(skill_id=sid, qtype=qtype, prompt=prompt, options_json=options_json, answer_json=answer_json, meta_json=meta_json)
        db.session.add(q)
        created += 1

    db.session.commit()
    flash(f"Imported. Created: {created}, Skipped: {skipped}.", "ok")
    return redirect(url_for("chairman.question_tool"))
